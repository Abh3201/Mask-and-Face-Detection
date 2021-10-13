import streamlit as st
from PIL import Image
import pandas as pd
import train
import datetime
import xlsxwriter 
import os
from openpyxl import load_workbook

Today_date = str(datetime.datetime.now())
Today_date = Today_date[:Today_date.index(' ')] + '.xlsx'
if Today_date not in os.listdir('ExcelData/'):
    workbook = xlsxwriter.Workbook('ExcelData/' + Today_date) 
    worksheet = workbook.add_worksheet()  
    worksheet.write('A1', 'Date') 
    worksheet.write('B1', 'Time') 
    worksheet.write('C1', 'Name') 
    worksheet.write('D1', 'NamePercentage') 
    worksheet.write('E1', 'Mask') 
    worksheet.write('F1', 'MaskPercentage')
    worksheet.write('G1', 'Temperature') 
    workbook.close() 

st.title('Live Stream!!')
FRAME_WINDOW = st.image([])

st.sidebar.title('Please select an Option!')
def main():
    a = st.sidebar.selectbox('Select',["Video", "Train new face","Retrain Model","Unknown", "Backup", "Exit"])
    print('A value is',a)
    if a == "Video":
        call_video()
    elif a == "Train new face":
        cv2.destroyAllWindows()
        name()
    elif a == "Retrain Model":
        cv2.destroyAllWindows()
        train.train()
        st.success('Model Trained Succesfully!!')
    elif a == "Unknown":
        cv2.destroyAllWindows()
        for x in os.listdir("Unknown"):
            st.image('Unknown/'+x)
    elif a == "Backup":
        cv2.destroyAllWindows()
        backup()
    elif a == "Exit":
        pass
# -*- coding: utf-8 -*-
# USAGE
# python detect_mask_video.py

# import the necessary packages
from tensorflow.keras.applications.mobilenet_v2 import preprocess_input
from tensorflow.keras.preprocessing.image import img_to_array
from tensorflow.keras.models import load_model
from imutils.video import VideoStream
import numpy as np
import argparse
import imutils
import time
import cv2
import os
import cv2
import numpy as np
import boto3
import shutil
import pymongo

myclient = pymongo.MongoClient("mongodb://localhost:27017/")
mydb = myclient["deepblue"]
mycol = mydb["user"]


recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('trainer/trainer.yml')
cascadePath = "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath);

font = cv2.FONT_HERSHEY_SIMPLEX

#iniciate id counter
id = 0

# names related to ids: example ==> Marcelo: id=1,  etc
names = ['Unknown']

def dbase():
    global names
    ls = ['Unknown']
    for x in mycol.find():
        ls.append(x['Name'])
    names = ls
    print(names)

# Initialize and start realtime video capture


# def train():
#     path = 'dataset'
    
#     recognizer = cv2.face.LBPHFaceRecognizer_create()
#     detector = cv2.CascadeClassifier("haarcascade_frontalface_default.xml");
    
#     # function to get the images and label data
#     def getImagesAndLabels(path):
    
#         imagePaths = [os.path.join(path,f) for f in os.listdir(path)]     
#         faceSamples=[]
#         ids = []
    
#         for imagePath in imagePaths:
    
#             PIL_img = Image.open(imagePath).convert('L') # convert it to grayscale
#             img_numpy = np.array(PIL_img,'uint8')
    
#             id = int(os.path.split(imagePath)[-1].split(".")[1])
#             faces = detector.detectMultiScale(img_numpy)
    
#             for (x,y,w,h) in faces:
#                 faceSamples.append(img_numpy[y:y+h,x:x+w])
#                 ids.append(id)
    
#         return faceSamples,ids
    
#     st.text("\n [INFO] Training faces. It will take a few seconds. Wait ...")
#     faces,ids = getImagesAndLabels(path)
#     recognizer.train(faces, np.array(ids))
    
#     # Save the model into trainer/trainer.yml
#     recognizer.write('trainer/trainer.yml') # recognizer.save() worked on Mac, but not on Pi
    
#     # Print the numer of faces trained and end program
#     print("\n [INFO] {0} faces trained. Exiting Program".format(len(np.unique(ids))))

#     st.text("\n [INFO] {0} faces trained. Exiting Program".format(len(np.unique(ids))))

def backup():
    s3_client = boto3.client('s3',
        aws_access_key_id='YOUR_KEY',
          aws_secret_access_key= 'YOUR_KEY')

    for file in os.listdir('Backup'):
            upload_file_bucket = 'deepblue'
            a = str(file)
            upload_file_key = 'User/'+ a[5:5+a[5:].index('.')] + "/" + str(a)
            s3_client.upload_file('Backup/' + file, upload_file_bucket, upload_file_key)
    shutil.rmtree("Backup") 
    os.mkdir("Backup")
    st.success('Backup Succesfully Completed!')

def data(name):
    cam = cv2.VideoCapture(0)
    cam.set(3, 640) # set video width
    cam.set(4, 480) # set video height
    # collection.find_one(sort=[("myfield", 1)])["myfield"]
    face_id = int(mycol.find_one(sort=[("_id", -1)])["_id"])
    face_id += 1

    

    # for x in mycol.find_one(sort = [("")]):
    #     print(int(x['_id']))
    face_detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
    
    # For each person, enter one numeric face id
    # face_id = input('\n enter user id end press <return> ==>  ')
    # face_id = st.text_input("Enter User ID :-")
    # name = st.text_input("Enter Name:-")
    # names.append(name)
    print("\n [INFO] Initializing face capture. Look the camera and wait ...")
    # Initialize individual sampling face count
    count = 0
    
    while(True):
    
        ret, img = cam.read()
      #  img = cv2.flip(img, -1) # flip video image vertically
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_detector.detectMultiScale(gray, 1.3, 5)
    
        for (x,y,w,h) in faces:
    
            cv2.rectangle(img, (x,y), (x+w,y+h), (255,0,0), 2)     
            count += 1
    
            # Save the captured image into the datasets folder
            cv2.imwrite("dataset/User." + str(face_id) + '.' + str(count)  + ".jpg", gray[y:y+h,x:x+w])
            cv2.imwrite("Backup/User." + str(face_id) + '.' + str(count) + ".jpg", gray[y:y+h,x:x+w])
    
            # cv2.imshow('image', img)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        FRAME_WINDOW.image(img)
    
        k = cv2.waitKey(100) & 0xff # Press 'ESC' for exiting video
        if k == 27:
            break
        elif count >= 30: # Take 30 face sample and stop video
             cv2.destroyAllWindows()
             cam.release()
             break
    
    # Do a bit of cleanup
    print("\n [INFO] Exiting Program and cleanup stuff")
    FRAME_WINDOW.image([])
    train.train()
    mydict = { "_id": face_id, "Name": name }

    x = mycol.insert_one(mydict)
    face_id = ""
    name = ""
    backup()
    st.success('Model Trained Succesfully!!')
    return
    
def name():
    # face_id = st.text_input("Enter User ID :-")
    name = st.text_input("Enter Name:-")
    if st.button('Start'):
        data(name)
        print("PROCESS COMPLETED")
    

def call(cam, minW, minH):

    ret, img =cam.read()
  #  img = cv2.flip(img, -1) # Flip vertically

    gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)

    faces = faceCascade.detectMultiScale( 
        gray,
        scaleFactor = 1.2,
        minNeighbors = 5,
        minSize = (int(minW), int(minH)),
       )
    id = None
    confidence = None
    for(x,y,w,h) in faces:

        cv2.rectangle(img, (x,y), (x+w,y+h), (0,255,0), 2)

        id, confidence = recognizer.predict(gray[y:y+h,x:x+w])

        # Check if confidence is less them 100 ==> "0" is perfect match 
        if (confidence < 100):
            id = names[id]
            confidence = "  {0}%".format(round(100 - confidence))
        else:
            id = "unknown"
            confidence = "  {0}%".format(round(100 - confidence))
        # cv2.putText(img, str(id), (x,y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 255, 0), 2)
        # cv2.putText(img, str(confidence), (x,y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 255, 0), 2)  
#         cv2.putText(frame, label, (startX, startY - 10),cv2.FONT_HERSHEY_SIMPLEX, 0.45, color, 2)
# 		cv2.rectangle(frame, (startX, startY), (endX, endY), color, 2)
    # cv2.imshow('camera',img) 
    
    return id, confidence

# Do a bit of cleanup
# print("\n [INFO] Exiting Program and cleanup stuff")
# cam.release()
# cv2.destroyAllWindows()




def detect_and_predict_mask(frame, faceNet, maskNet, args):
	# grab the dimensions of the frame and then construct a blob
	# from it
	(h, w) = frame.shape[:2]
	blob = cv2.dnn.blobFromImage(frame, 1.0, (300, 300),
		(104.0, 177.0, 123.0))

	# pass the blob through the network and obtain the face detections
	faceNet.setInput(blob)
	detections = faceNet.forward()

	# initialize our list of faces, their corresponding locations,
	# and the list of predictions from our face mask network
	faces = []
	locs = []
	preds = []

	# loop over the detections
	for i in range(0, detections.shape[2]):
		# extract the confidence (i.e., probability) associated with
		# the detection
		confidence = detections[0, 0, i, 2]

		# filter out weak detections by ensuring the confidence is
		# greater than the minimum confidence
		if confidence > args["confidence"]:
			# compute the (x, y)-coordinates of the bounding box for
			# the object
			box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
			(startX, startY, endX, endY) = box.astype("int")

			# ensure the bounding boxes fall within the dimensions of
			# the frame
			(startX, startY) = (max(0, startX), max(0, startY))
			(endX, endY) = (min(w - 1, endX), min(h - 1, endY))

			# extract the face ROI, convert it from BGR to RGB channel
			# ordering, resize it to 224x224, and preprocess it
			face = frame[startY:endY, startX:endX]
			face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
			face = cv2.resize(face, (300, 300))
			face = img_to_array(face)
			face = preprocess_input(face)

			# add the face and bounding boxes to their respective
			# lists
			faces.append(face)
			locs.append((startX, startY, endX, endY))

	# only make a predictions if at least one face was detected
	if len(faces) > 0:
		# for faster inference we'll make batch predictions on *all*
		# faces at the same time rather than one-by-one predictions
		# in the above `for` loop
		faces = np.array(faces, dtype="float32")
		preds = maskNet.predict(faces, batch_size=32)

	# return a 2-tuple of the face locations and their corresponding
	# locations
	return (locs, preds)

# construct the argument parser and parse the arguments

# vs = VideoStream(src=0).start()
# time.sleep(2.0)

# loop over the frames from the video stream
def call_video():
    dbase()
    cam = cv2.VideoCapture(0)
    cam.set(3, 640) # set video widht
    cam.set(4, 480) # set video height
    
    # Define min window size to be recognized as a face
    minW = 0.1*cam.get(3)
    minH = 0.1*cam.get(4)
    ap = argparse.ArgumentParser()
    ap.add_argument("-f", "--face", type=str,
    	default="face_detector",
    	help="path to face detector model directory")
    ap.add_argument("-m", "--model", type=str,
    	default="mask_detector.model",
    	help="path to trained face mask detector model")
    ap.add_argument("-c", "--confidence", type=float, default=0.5,
    	help="minimum probability to filter weak detections")
    args = vars(ap.parse_args())
    
    # load our serialized face detector model from disk
    print("[INFO] loading face detector model...")
    prototxtPath = os.path.sep.join([args["face"], "deploy.prototxt"])
    weightsPath = os.path.sep.join([args["face"],
    	"res10_300x300_ssd_iter_140000.caffemodel"])
    faceNet = cv2.dnn.readNet(prototxtPath, weightsPath)
    
    # load the face mask detector model from disk
    print("[INFO] loading face mask detector model...")
    maskNet = load_model(args["model"])
    
    # initialize the video stream and allow the camera sensor to warm up
    print("[INFO] starting video stream...")
    while True:
    	# grab the frame from the threaded video stream and resize it
    	# to have a maximum width of 400 pixels
    	_, frame = cam.read()
    # 	frame = imutils.resize(frame, width=300)
    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    	frame1 = frame.copy()

    	# detect faces in the frame and determine if they are wearing a
    	# face mask or not
    	(locs, preds) = detect_and_predict_mask(frame, faceNet, maskNet, args)
    
    	# loop over the detected face locations and their corresponding
    	# locations
    	Cu_date = Time = Name = Nameper = Mask = Maskper = None
    	for (box, pred) in zip(locs, preds):
    		# unpack the bounding box and predictions
    		(startX, startY, endX, endY) = box
    		(mask, withoutMask) = pred
    
    		# determine the class label and color we'll use to draw
    		# the bounding box and text
    		label = "Mask" if mask > withoutMask else "No Mask"
    		color = (0, 255, 0) if label == "Mask" else (0, 0, 255)
    
    		# include the probability in the label
    		label = "{}: {:.2f}%".format(label, max(mask, withoutMask) * 100)
    
    		# display the label and bounding box rectangle on the output
    		# frame
    		cv2.putText(frame, label, (startX, startY - 10),
    			cv2.FONT_HERSHEY_SIMPLEX, 0.45, color, 2)
    		Cu_date = str(datetime.datetime.now())
    		Cu_date, Time = Cu_date[:Cu_date.index(' ')], Cu_date[Cu_date.index(' ')+1:]
    		# print('Lenth of Date', len([Cu_date]))
    		Mask, Maskper = label[:label.index(':')], label[label.index(':')+1:]
           
    		img1, img2 = call(cam, minW, minH)
    		# print(img1, img2)
    		if img1 != None and img2 != None:
    		    label = "{}:{}".format(str(img1), str(img2))
    		    cv2.putText(frame, label, (startX, startY - 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.45, color, 2)
    		# cv2.putText(frame, img2, (startX + 10, startY - 30),
    		# 	cv2.FONT_HERSHEY_SIMPLEX, 0.45, color, 2)
    		Name, Nameper = img1, img2
    		cv2.rectangle(frame, (startX, startY), (endX, endY), color, 2)
    	if Name == 'unknown':
            n = len(os.listdir("Unknown")) 
            cv2.imwrite("Unknown/unknown." + str(n+1) + ".jpg",frame)
    	if Time != None and Name != None :
            df = pd.DataFrame({
                    'Date': [Cu_date],
                    'Time': [Time],
                    'Name': [Name],
                    'NamePercentage': [Nameper],
                    'Mask': [Mask],
                    'MaskPercentage': [Maskper],
                    'Temperature' : ['']})
            writer = pd.ExcelWriter('ExcelData/' + Today_date, engine='openpyxl')
            writer.book = load_workbook('ExcelData/' + Today_date)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            reader = pd.read_excel(r'ExcelData/' + Today_date,  engine='openpyxl')
            df.to_excel(writer,index=False,header=False,startrow=len(reader)+1)
            writer.close()

    	# show the output frame
    # 	cv2.imshow("Frame", frame)
    	frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    	FRAME_WINDOW.image(frame)
    # 	img = call()
    # 	img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    # 	FRAME_WINDOW.image(img)
    # 	frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        
        # frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        # FRAME_WINDOW.image(frame)  
    	key = cv2.waitKey(1) & 0xFF
    
    	# if the `q` key was pressed, break from the loop
    	if key == ord("q"):
    		break
    cv2.destroyAllWindows()
    cam.release()

# do a bit of cleanup
# cv2.destroyAllWindows()
# vs.stop()  ["Video", "Train new face", "Exit"]


if __name__ == "__main__":
    main()
